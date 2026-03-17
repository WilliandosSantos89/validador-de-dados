import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, date
import re
import json

COLUNAS_OBRIGATORIAS = [
    "ID", "Material", "Plataforma", "Responsável",
    "Prazo", "Status"
]

PLATAFORMAS_VALIDAS  = ["LMS Moodle", "Portal do Aluno", "App Arco"]
STATUS_VALIDOS       = ["Publicado", "Pendente", "Atrasado", "Erro"]
RESPONSAVEIS_VALIDOS = ["Ana Lima", "Carlos Souza", "Fernanda Reis", "Pedro Alves"]

REGRAS = {
    "ID":           lambda v: bool(re.match(r"^MAT-\d{3}$", str(v).strip())),
    "Material":     lambda v: len(str(v).strip()) >= 5,
    "Plataforma":   lambda v: str(v).strip() in PLATAFORMAS_VALIDAS,
    "Responsável":  lambda v: str(v).strip() in RESPONSAVEIS_VALIDOS,
    "Prazo":        lambda v: validar_data(v),
    "Status":       lambda v: str(v).strip() in STATUS_VALIDOS,
}

MENSAGENS = {
    "ID":          "ID fora do padrão MAT-000",
    "Material":    "Nome do material muito curto ou vazio",
    "Plataforma":  f"Plataforma inválida — use: {', '.join(PLATAFORMAS_VALIDAS)}",
    "Responsável": f"Responsável não cadastrado",
    "Prazo":       "Data inválida — use o formato DD/MM/YYYY",
    "Status":      f"Status inválido — use: {', '.join(STATUS_VALIDOS)}",
}


def validar_data(valor):
    if not valor or str(valor).strip() in ("—", ""):
        return True
    try:
        datetime.strptime(str(valor).strip(), "%d/%m/%Y")
        return True
    except ValueError:
        return False


def ler_planilha(arquivo):
    wb   = openpyxl.load_workbook(arquivo)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return [], []

    cabecalho = [str(c).strip() if c else "" for c in rows[0]]
    dados     = []

    for row in rows[1:]:
        if any(row):
            linha = {}
            for i, col in enumerate(cabecalho):
                linha[col] = row[i] if i < len(row) else None
            dados.append(linha)

    return cabecalho, dados


def validar(cabecalho, dados):
    erros_cabecalho = []
    for col in COLUNAS_OBRIGATORIAS:
        if col not in cabecalho:
            erros_cabecalho.append(f"Coluna obrigatória ausente: '{col}'")

    resultados = []
    total_erros = 0

    for i, linha in enumerate(dados, start=2):
        erros_linha = []

        for campo, regra in REGRAS.items():
            valor = linha.get(campo)
            if valor is None or str(valor).strip() == "":
                erros_linha.append({
                    "campo":   campo,
                    "valor":   "",
                    "motivo":  "Campo obrigatório vazio",
                })
            elif not regra(valor):
                erros_linha.append({
                    "campo":  campo,
                    "valor":  str(valor).strip(),
                    "motivo": MENSAGENS[campo],
                })

        resultados.append({
            "linha":   i,
            "id":      linha.get("ID", "—"),
            "material": str(linha.get("Material", ""))[:40],
            "erros":   erros_linha,
            "valido":  len(erros_linha) == 0,
        })

        total_erros += len(erros_linha)

    return erros_cabecalho, resultados, total_erros


def gerar_relatorio(cabecalho, dados, erros_cabecalho, resultados, total_erros):
    wb  = openpyxl.Workbook()

    validos   = [r for r in resultados if r["valido"]]
    invalidos = [r for r in resultados if not r["valido"]]

    # ── ABA 1: RESUMO ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumo"

    ws1.merge_cells("A1:D1")
    t = ws1["A1"]
    t.value     = f"Relatório de Validação — {date.today().strftime('%d/%m/%Y')}"
    t.font      = Font(bold=True, size=13, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor="1B4332")
    t.alignment = Alignment(horizontal="center")
    ws1.row_dimensions[1].height = 28

    def linha_resumo(ws, row, label, valor, cor_fundo=None, cor_texto="000000"):
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=row, column=2, value=valor)
        c.alignment = Alignment(horizontal="center")
        if cor_fundo:
            c.fill = PatternFill("solid", fgColor=cor_fundo)
            c.font = Font(bold=True, color=cor_texto)

    linha_resumo(ws1, 3,  "Total de registros",   len(resultados))
    linha_resumo(ws1, 4,  "Registros válidos",     len(validos),   "C8E6C9", "2D6A4F")
    linha_resumo(ws1, 5,  "Registros com erro",    len(invalidos), "FFCDD2", "C62828")
    linha_resumo(ws1, 6,  "Total de erros",        total_erros,    "FFCCBC", "E64A19")
    linha_resumo(ws1, 7,  "Taxa de qualidade",
                 f"{round(len(validos)/len(resultados)*100, 1)}%" if resultados else "0%",
                 "E3F2FD", "1565C0")

    if erros_cabecalho:
        ws1.cell(row=9, column=1, value="Erros de estrutura:").font = Font(bold=True, color="C62828")
        for i, err in enumerate(erros_cabecalho, start=10):
            ws1.cell(row=i, column=1, value=err).font = Font(color="C62828")

    for col, larg in zip(["A","B","C","D"], [30, 20, 20, 20]):
        ws1.column_dimensions[col].width = larg

    # ── ABA 2: ERROS ─────────────────────────────────────────
    ws2 = wb.create_sheet("Erros")

    ws2.merge_cells("A1:E1")
    t2 = ws2["A1"]
    t2.value     = "Registros com Erro — Correção Necessária"
    t2.font      = Font(bold=True, size=12, color="FFFFFF")
    t2.fill      = PatternFill("solid", fgColor="B71C1C")
    t2.alignment = Alignment(horizontal="center")

    for col, texto in enumerate(["Linha", "ID", "Material", "Campo", "Motivo"], start=1):
        c = ws2.cell(row=2, column=col, value=texto)
        c.font      = Font(bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="1B4332")
        c.alignment = Alignment(horizontal="center")

    row = 3
    for r in invalidos:
        for erro in r["erros"]:
            ws2.cell(row=row, column=1, value=r["linha"])
            ws2.cell(row=row, column=2, value=r["id"])
            ws2.cell(row=row, column=3, value=r["material"])
            c4 = ws2.cell(row=row, column=4, value=erro["campo"])
            c5 = ws2.cell(row=row, column=5, value=erro["motivo"])
            c4.fill = PatternFill("solid", fgColor="FFCCBC")
            c5.fill = PatternFill("solid", fgColor="FFCDD2")
            for col in range(1, 6):
                ws2.cell(row=row, column=col).alignment = \
                    Alignment(horizontal="center")
            row += 1

    for col, larg in zip(["A","B","C","D","E"], [8, 12, 42, 18, 42]):
        ws2.column_dimensions[col].width = larg

    # ── ABA 3: VÁLIDOS ───────────────────────────────────────
    ws3 = wb.create_sheet("Válidos")

    ws3.merge_cells("A1:F1")
    t3 = ws3["A1"]
    t3.value     = "Registros Aprovados na Validação"
    t3.font      = Font(bold=True, size=12, color="FFFFFF")
    t3.fill      = PatternFill("solid", fgColor="1B4332")
    t3.alignment = Alignment(horizontal="center")

    for col, texto in enumerate(
        ["Linha", "ID", "Material", "Plataforma", "Responsável", "Status"], start=1):
        c = ws3.cell(row=2, column=col, value=texto)
        c.font      = Font(bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="2D6A4F")
        c.alignment = Alignment(horizontal="center")

    for i, r in enumerate(validos, start=3):
        linha = dados[r["linha"] - 2]
        valores = [r["linha"], r["id"], r["material"],
                   linha.get("Plataforma",""), linha.get("Responsável",""),
                   linha.get("Status","")]
        for col, val in enumerate(valores, start=1):
            c = ws3.cell(row=i, column=col, value=val)
            c.alignment = Alignment(horizontal="center")
            c.fill = PatternFill("solid", fgColor="C8E6C9")

    for col, larg in zip(["A","B","C","D","E","F"], [8, 12, 42, 18, 18, 12]):
        ws3.column_dimensions[col].width = larg

    nome = f"validacao_{date.today().strftime('%Y%m%d')}.xlsx"
    wb.save(nome)

    # log JSON
    log = {
        "data":           date.today().isoformat(),
        "total":          len(resultados),
        "validos":        len(validos),
        "invalidos":      len(invalidos),
        "total_erros":    total_erros,
        "taxa_qualidade": f"{round(len(validos)/len(resultados)*100,1)}%" if resultados else "0%",
        "registros":      resultados,
    }
    with open(f"log_validacao_{date.today().strftime('%Y%m%d')}.json", "w",
              encoding="utf-8") as f:
        json.dump(log, f, ensure_ascii=False, indent=2)

    return nome


def main():
    arquivo = "materiais.xlsx"

    print(f"Lendo planilha: {arquivo}")
    cabecalho, dados = ler_planilha(arquivo)

    if not dados:
        print("Planilha vazia ou sem dados.")
        return

    print(f"Validando {len(dados)} registro(s)...")
    erros_cab, resultados, total_erros = validar(cabecalho, dados)

    validos   = [r for r in resultados if r["valido"]]
    invalidos = [r for r in resultados if not r["valido"]]

    nome = gerar_relatorio(cabecalho, dados, erros_cab, resultados, total_erros)

    print(f"\nValidação concluída.")
    print(f"Total          : {len(resultados)}")
    print(f"Válidos        : {len(validos)}")
    print(f"Com erro       : {len(invalidos)}")
    print(f"Total de erros : {total_erros}")
    print(f"Taxa qualidade : {round(len(validos)/len(resultados)*100,1)}%")
    print(f"\nRelatório salvo: {nome}")

    if invalidos:
        print(f"\nRegistros com problema:")
        for r in invalidos:
            for e in r["erros"]:
                print(f"  Linha {r['linha']} | {r['id']} | {e['campo']} — {e['motivo']}")


if __name__ == "__main__":
    main()